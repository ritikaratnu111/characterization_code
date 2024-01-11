import os
import logging
from openpyxl import Workbook
from openpyxl import load_workbook

class Predictor():
	def __init__(self, tb, assembly):     
		self.tb = tb
		assembly.predictor = {}
		self.wb = load_workbook(filename = f"{self.tb}/characterization.xlsx")
		self.sheet = self.wb.active
		self.data = {}

	def read_db(self):
		for row in sheet.iter_rows(min_row=2, values_only=True):
			experiment_name = row[0]
			components_data = []

		    # Iterate through the rest of the row data
		    for i in range(1, len(row), len(power_types) * len(sub_types)):
		        component_name = component_names[(i - 1) // (len(power_types) * len(sub_types))]
		        component_powers = []

		        # Iterate through power types and subtypes
		        for j in range(i, i + len(power_types) * len(sub_types), len(sub_types)):
		            power_type = power_types[(j - i) // len(sub_types)]
		            power_subtype = sub_types[(j - i) % len(sub_types)]
		            power_value = row[j]
		            component_powers.append({
		                "type": power_type,
		                "subtype": power_subtype,
		                "value": power_value
		            })

		        components_data.append({
		            "component": component_name,
		            "powers": component_powers
		        })

		    self.data.append({
		        "experiment": experiment_name,
		        "components": components_data
		    })

		# Print the read data
		for experiment_data in data:
		    print("Experiment:", experiment_data["experiment"])
		    for component_data in experiment_data["components"]:
		        print("Component:", component_data["component"])
		        for power_data in component_data["powers"]:
		            print(f"{power_data['type']} - {power_data['subtype']}: {power_data['value']}")
		    print()

	def get_active_component_active_energy(self):
		print("Getting active component active energy")
		#Multiply per cycle power from self.data with active cycles from assembly and write in assembly.cells[id]['energy']['predictor']['active_components'][component]['internal']['active']

	def get_active_component_inactive_energy(self):
		print("Getting active component inactive energy")
		#Multiply per cycle power from self.data with inactive cycles from assembly and write in assembly.cells[id]['energy']['predictor']['active_components'][component]['internal']['inactive']

	def get_inactive_component_energy(self):
		print("Getting inactive component energy")
		#Multiply per cycle power from self.data with total cycles from assembly and write in assembly.cells[id]['energy']['predictor']['inactive_components'][component]['internal']['inactive']

	def get_total_energy(self):
		print("Getting total energy")
		#Add active and inactive energy from assembly and write in assembly.cells[id]['energy']['predictor']['total']
		