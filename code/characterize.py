import os
import logging
import random
import constants
from openpyxl import Workbook
from simulation import Simulation
from innovus_reader import InnovusPowerParser
from measurement import Measurement
import json
import matplotlib.pyplot as plt
import numpy as np

class Characterize():
	
    def __init__(self, tb, vcd_dir, cells):     
        self.tb = tb
        self.vcd_dir = vcd_dir
        os.environ['VCD_DIR'] = vcd_dir
        os.chdir(tb)
        os.makedirs(vcd_dir, exist_ok=True)
        self.cells = cells
        self.count = 0
        self.reader = InnovusPowerParser()

    def run_simulation(self, window, i, uid):
        start = window['start']
        end = window['end']
        Simulation.trigger_vsim(i, start, end, False, uid)
        Simulation.trigger_innovus(i, start, end, False)
        Simulation.remove_file(f"{self.tb}/{self.vcd_dir}/iter_{i}.vcd")
        Simulation.remove_file(f"{self.tb}/testbench_rtl_{uid}.vhd")
        Simulation.remove_dir(f"{self.tb}/work_{uid}")

    def generate_randomized_mem_init_files(self, start, end):
        Simulation.generate_randomized_mem_init_files(start, end)

    def run_randomized_simulation(self, start, end):
        print("Running randomized simulation")
        Simulation.generate_randomized_mem_init_files(start, end)
        for i in range(start, end):
            print(i)
            uid = Simulation.update_mem_init_file(self.tb, i)
            print(uid)
            for cell in self.cells:
                self.run_simulation(cell.total_window, i, uid)
                #cell.get_measurement(self.reader,cell.tiles,i,self.tb)
    
    def get_average(self,start, end):
        for cell in self.cells:
            cell.profiler.set_avg_measurement_size(end-start)
            hist_data = []
            for i in range(start,end):
                file_path = f"{self.tb}/vcd/iter_{i}.json"
                with open(file_path, 'r') as file:
                    data = json.load(file)
                    cell_data = data[cell.cell_id]
                    measurement = Measurement()
                    measurement.power.internal = cell_data.get('power', {}).get('internal', 0)
                    measurement.power.switching = cell_data.get('power', {}).get('switching', 0)
                    measurement.power.leakage = cell_data.get('power', {}).get('leakage', 0)
                    hist_data.append(measurement.power.switching)
                    print(measurement.power.internal, measurement.power.switching, measurement.power.leakage)
                    cycles = int(cell_data.get('cycles', 0))
                    measurement.power.total = measurement.power.internal + measurement.power.switching + measurement.power.leakage
                    measurement.set_window({'start': 0, 'end': cycles, 'clock_cycles': cycles})
                    measurement.get_energy()
                    cell.profiler.add_avg_measurement(measurement, i)
                    cell.profiler.print_avg_results(i)
            
            #calculate std of hist
#            std_list = []
#            print("hello")
#            print(hist_data)
#            for i in range(len(hist_data)-2):
#                slice = hist_data[:i+2]
#                std = np.std(slice)
#                std_list.append(std)
#                print(std)
#            print(std_list)
#            plt.figure()
#            plt.plot(std_list)
#            plt.savefig('std_switching_energy_plot.png')


            cell.profiler.plot()
#            plt.hist(hist_data, bins=20)
#            plt.savefig('hist_switching_energy_plot.png')

    def get_cell_measurements(self,start,end):
        for i in range(start,end):
            pwr_file=f"{self.tb}/vcd/iter_{i}.vcd.pwr"
            json_file=f"{self.tb}/vcd/iter_{i}.json"
            print(pwr_file,json_file)
            if os.path.exists(pwr_file) and not os.path.exists(json_file):
                total_nets = 0
                balance = 0
                accounted_nets = 0
                self.reader.update_nets(pwr_file)
                for cell in self.cells:
                    cell.profiler.set_measurement(self.reader,cell.tiles, cell.components.active)
                    total_nets = cell.profiler.measurement_set.actual.nets
                    balance = total_nets
                    print(cell.cell_id,cell.tiles,cell.total_window)
                    print(balance, (balance / total_nets) * 100)
                    self.reader.remove_labels(cell.tiles)
                    for component in cell.components.active:
                        self.reader.get_count_of_inactive_labels(cell.tiles)
                        #if(component.name != "dpu"):
                        #    continue
                        logging.info('%s %s %s',component.name, component.signals, component.active_window)
                        component.profiler.set_measurement(self.reader)
                        accounted_nets += component.profiler.measurement.actual.nets
                        balance = total_nets - accounted_nets
                        logging.info('%s %s %s',component.profiler.measurement.actual.nets, balance, (balance / total_nets) * 100)
                        T = cell.total_window['clock_cycles']
                        cell.profiler.measurement_set.active.add(
                            T, T, T,
                            cell.profiler.measurement_set.active, 
                            component.profiler.measurement.predicted)
                        cell.profiler.measurement_set.set_predicted()
                        cell.profiler.measurement_set.set_error()
                        cell.profiler.measurement_set.error.log()
                self.write_json(i)

    def write_json(self,i):
        data = {}
        for cell in self.cells:
            for component in cell.components.active:
                #if(component.name != "dpu"):
                #    continue
                #Write the component, active cycles of the component, and power of all three types in a json file in the same directory
                #as the testbench
                data[component.name] = {
                    "active": {
                        "window": f"{component.profiler.active_window['windows']}" if component.active_window else "None",
                        "cycles": component.profiler.active_window['clock_cycles'] if component.active_window else 0,
                        "power": {
                            "internal": component.profiler.measurement.active.power.internal,
                            "switching": component.profiler.measurement.active.power.switching,
                            "leakage": component.profiler.measurement.active.power.leakage
                        }
                    },
                    "inactive": {
                        "window": f"{component.profiler.inactive_window['windows']}" if component.inactive_window else "None",
                        "cycles": component.profiler.inactive_window['clock_cycles'] if component.inactive_window else 0,
                        "power": {
                            "internal": component.profiler.measurement.inactive.power.internal,
                            "switching": component.profiler.measurement.inactive.power.switching,
                            "leakage": component.profiler.measurement.inactive.power.leakage
                        }
                    }
                }
            data[cell.cell_id] = {
                "window": f"{cell.total_window}",
                "cycles": cell.total_window['clock_cycles'],
                "power": {
                    "internal": cell.profiler.measurement.actual.power.internal,
                    "switching": cell.profiler.measurement.actual.power.switching,
                    "leakage": cell.profiler.measurement.actual.power.leakage
                }
            }
        with open(f"{self.tb}/vcd/iter_{i}.json", "w") as file:
            json.dump(data, file, indent=2)


#            for component in cell.components.active:
#                component.profiler.set_avg_measurement_size(num_iterations)
#                for i in range(num_iterations):
#                    file_path = f"{self.tb}/vcd/iter_{i}.json"
#                    with open(file_path, 'r') as file:
#                        data = json.load(file)
#                    component_data = data.get(component.name, {})
#                    active_measurement = Measurement()
#                    inactive_measurement = Measurement()
#                    if 'active' in component_data:
#                        active_measurement.power.internal = component_data.get('active', {}).get('power', {}).get('internal', 0)
#                        active_measurement.power.switching = component_data.get('active', {}).get('power', {}).get('switching', 0)
#                        active_measurement.power.leakage = component_data.get('active', {}).get('power', {}).get('leakage', 0)
#                        active_measurement.power.total = active_measurement.power.internal + active_measurement.power.switching + active_measurement.power.leakage
#                        cycles = int(component_data.get('active', {}).get('cycles', 0))
#                        active_measurement.set_window({'start': 0, 'end': cycles * constants.CLOCK_PERIOD, 'clock_cycles': cycles})
#                        active_measurement.get_energy()
#
#                    if 'inactive' in component_data:
#                        inactive_measurement.power.internal = component_data.get('inactive', {}).get('power', {}).get('internal', 0)
#                        inactive_measurement.power.switching = component_data.get('inactive', {}).get('power', {}).get('switching', 0)
#                        inactive_measurement.power.leakage = component_data.get('inactive', {}).get('power', {}).get('leakage', 0)
#                        inactive_measurement.power.total = inactive_measurement.power.internal + inactive_measurement.power.switching + inactive_measurement.power.leakage
#                        cycles = int(component_data.get('inactive', {}).get('cycles', 0))
#                        inactive_measurement.set_window({'start': 0, 'end': cycles * constants.CLOCK_PERIOD, 'clock_cycles': cycles})
#                    component.profiler.add_avg_measurement(active_measurement, inactive_measurement, i)
#                    if(component.name == "noc"):
#                        component.profiler.print_avg_results(i)

                    
#
#    def check_net_balance(self, count):
#        for i in range(count):
#            total_nets = 0
#            balance = 0
#            accounted_nets = 0
#            if i == 0:
#                for cell in self.cells:
#                    total_nets = cell.profiler.nets
#                    for component in cell.components.active:
#                        accounted_nets += component.profiler.nets
#                        balance = total_nets - accounted_nets
#                        print(balance, (balance / total_nets) * 100)
#                    for component in cell.components.inactive:
#                        print(component.name, component.signals)
#                        accounted_nets += component.profiler.nets
#                        balance = total_nets - accounted_nets
#                        print(balance, (balance / total_nets) * 100)
#            else:
#                continue
#
#    def check_energy_balance(self,count):
#        for i in range(count):
#            if (i == 0):
#                total_measurement = Measurement()
#                active_measurement = Measurement()
#                inactive_measurement = Measurement()
#                balance_measurement = Measurement()
#                for cell in self.cells:
#                    total_measurement = cell.profiler.total_measurement
#                    for component in cell.components.active:
#                        for measurement in component.profiler.active_measurement:
#                            active_measurement.add_energy(measurement)
#                            balance_measurement.diff_energy(total_measurement,active_measurement)
#                            balance_measurement.log_energy()
#                        for measurement in component.profiler.inactive_measurement:
#                            active_measurement.add_energy(measurement)
#                            balance_measurement.diff_energy(total_measurement,active_measurement)
#                            balance_measurement.log_energy()
#                    for component in cell.components.inactive:
#                        inactive_measurement.add_energy(component.profiler.inactive_measurement)
#                        balance_measurement.diff_energy(total_measurement,inactive_measurement)
#                        balance_measurement.log_energy()
#
#
#    def get_AEC_measurement(self):
#        print("Getting active component measurement")
#        for cell in self.cells:
#            for component in cell.components.active:
#                print(component.name)
#                logging.info("Setting active measurement for %s", component.name)
#                component.set_active_measurement(self.reader, 0)
#                logging.info("Setting inactive measurement for %s", component.name)
#                component.set_inactive_measurement(self.reader, 0)
#
#    def write_db(self):
#        print("Writing to database")
#        data = {
#            "Experiment1": {
#                "Component1": {
#                    "active": {"internal": 1, "switching": 1, "leakage": 1},
#                    "inactive": {"internal": 0, "switching": 0, "leakage": 0}
#                },
#                "Component2": {
#                    "active": {"internal": 2, "switching": 2, "leakage": 2},
#                    "inactive": {"internal": 0, "switching": 0, "leakage": 0}
#                },
#                "Component3": {
#                    "active": {"internal": 3, "switching": 3, "leakage": 3},
#                    "inactive": {"internal": 6, "switching": 4, "leakage": 3}
#                }
#            },
#            # Add more experiments if needed
#        }
#    
#        components = ["Component1", "Component2", "Component3"]  
#        states = ["active", "inactive"]
#        power_types = ["internal", "switching", "leakage"]
#    
#        wb = Workbook()
#        sheet = wb.active
#        print("Writing to database")
#        data = {
#                    "Experiment1": {
#                        "Component1": {
#                            "active": {"internal": 1, "switching": 1, "leakage": 1},
#                            "inactive": {"internal": 0, "switching": 0, "leakage": 0}
#                        },
#                        "Component2": {
#                            "active": {"internal": 2, "switching": 2, "leakage": 2},
#                            "inactive": {"internal": 0, "switching": 0, "leakage": 0}
#                        },
#                        "Component3": {
#                            "active": {"internal": 3, "switching": 3, "leakage": 3},
#                            "inactive": {"internal": 6, "switching": 4, "leakage": 3}
#                        }
#                    },
#                    # Add more experiments if needed
#                }
#
#        components = ["Component1", "Component2", "Component3"]  
#        states = ["active", "inactive"]
#        power_types = ["internal", "switching", "leakage"]
#    
#        wb = Workbook()
#        sheet = wb.active
#    
#        no_of_components = 3
#        sub_cols = 3 * 2
#        no_of_cols = no_of_components * sub_cols
#    
#        # Write headers
#        headers = ["Experiments"]
#        headers.extend([""] * no_of_cols)
#        sheet.append(headers)
#    
#        sub_headers = [""]
#        for component in components:
#            sub_headers.append(component)
#            sub_headers.extend([""] * (sub_cols - 1))  
#        sheet.append(sub_headers)
#
#        sub_sub_headers = [""]
#        for state in states:
#            sub_headers.append(state)
#            sub_headers.extend([""] * 2)
#        sheet.append(sub_sub_headers)
#
#        sub_sub_sub_headers = [""]
#        sub_sub_sub_headers.extend(power_types * no_of_components)
#        sheet.append(sub_sub_sub_headers)
#    
#        # Write data to the sheet
#        for exp, info in data.items():
#            row = [exp]
#            for component in components:
#                for state in states:
#                    powers = info[component][state]
#                    row.extend([powers.get("internal", ""), powers.get("switching", ""), powers.get("leakage", "")])
#            sheet.append(row)
#    
#        # Merge cells for sub-columns
#        #sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=10)
#    
#        # Save the workbook
#        wb.save("experiment_data.xlsx")
#
#    def run_simulation_per_component(self, i):
#        for cell in self.cells:
#            for component in cell.components.active:
#                print(component.name)
#                if component.name == "bus_selector" or component.name == "dimarch_data_in_out_reg":
#                    tile = component.signals[0].split('*')[0]
#                    for window in component.active_window: 
#                        state = "active"
#                        Simulation.trigger_vsim(i, window['start'], window['end'], False, True, state, tile, component.name)
#                        Simulation.trigger_innovus(i, window['start'], window['end'], False, True, state, tile, component.name)
#                    for window in component.inactive_window: 
#                        state = "inactive"
#                        Simulation.trigger_vsim(i, window['start'], window['end'], False, True, state, tile, component.name)
#                        Simulation.trigger_innovus(i, window['start'], window['end'], False, True, state, tile, component.name)
#
    def run_simulation_per_cycle(self):
        uid = "per_cycle"
        window = self.cells[0].total_window
        Simulation.trigger_vsim(0, window['start'], window['end'],  True, uid)
        Simulation.trigger_innovus(0, window['start'], window['end'],  True)
#
    def get_per_cycle_measurement(self):
            print("Getting per cycle measurement")
            # Add an indented block here
            for cell in self.cells:
                logging.info("Setting per_cycle measurement for %s", cell.drra_tile)
                for component in cell.components.active:
                    # if(component.name != "noc"):
                    #     continue
                    print(component.name)
                    logging.info("Setting per_cycle measurement for %s", component.name)
                    component.profiler.set_per_cycle_measurement(self.reader, component.signals)
            for component in cell.components.inactive:
                # if(component.name != "data_selector"):
                #     continue
                print(component.name, component.signals)
                logging.info("Setting per_cycle measurement for %s", component.name)
                component.profiler.set_per_cycle_measurement(self.reader, component.signals)
#
#    def get_AEC_measurements_from_per_cycle(self):
#        print("Getting per cycle component measurement")
#        for cell in self.cells:
#            logging.info("Setting component measurement from per_cycle for %s", cell.drra_tile)
#            for component in cell.components.active:
#                # if(component.name != "noc"):
#                #     continue
#                print(component.name)
#                logging.info("Setting active measurement from per_cycle for %s", component.name)
#                component.profiler.set_active_measurement_from_per_cycle()
#                component.profiler.set_inactive_measurement_from_per_cycle()
#                component.profiler.set_total_measurement_from_per_cycle(cell.total_window)
#                component.profiler.set_total_measurement_from_iter_file(self.reader, component.signals, 0, cell.total_window)
#
